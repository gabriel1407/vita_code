import datetime
from datetime import datetime
import tempfile
import os
import requests
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from PIL import Image
from django.http import JsonResponse
from django.conf import settings
from datetime import timedelta
from django.shortcuts import render, redirect
#from django.contrib.auth.models import User
from django.contrib.auth import get_user_model
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse, Http404
from django.utils.timezone import localtime
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from rest_framework import status, serializers
from rest_framework import viewsets
from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import action, api_view
from rest_framework.generics import GenericAPIView
from rest_framework.viewsets import ModelViewSet, ReadOnlyModelViewSet
from rest_framework.parsers import JSONParser, FormParser
from rest_framework.response import Response
from rest_framework.serializers import Serializer
from rest_framework.views import APIView
from task.models import Task, TaskHistory, project_tasks
from task.serializers import TaskListSerializer, TaskSerializer, TaskHistoryListSerializer, project_tasks_serializer, project_tasks_list_serializer
from task.filters import CreatedBetweenFilter
from companies.models import Company
from companies.serializers import UserListSerializer
from redmail import EmailSender
from django.template import loader
from users.models import UserCustomer
# Create your views here.

User = get_user_model()

class TaskViewSet(ModelViewSet):
    #permission_classes = (IsAppAuthenticated, IsAppStaff, IsAuthenticated, IsSuperUser)
    serializer_class = TaskListSerializer
    queryset = Task.objects.all()
    filter_backends = (DjangoFilterBackend,)
    filterset_fields = ('is_enabled', 'user', 'id')
    

    def send_email(self, request, id, *args, **kwargs):
        try:
            users = UserCustomer.objects.filter(id__in=request.data.get("user"))
            task_day = Task.objects.filter(id=id).first()

            for user in users:
                print("Usuario: ", user.email)
                if user is not None:
                    #image = Image.open('C:/Users/gabri/OneDrive/Documentos/Repositorios-github/sistema_gestion_actividades/Sistema_gestion_actividades/task/templates/actividades-de-trabajo-en-equipo.png')
                    image_path = f'{settings.STATIC_ROOT}/media/templates/actividades-de-trabajo-en-equipo.png'
                    image = Image.open(image_path)
                    #image = Image.open('C:/Users/gabri/Documents/GitHub/sistema_gestion_actividades/Sistema_gestion_actividades/task/templates/actividades-de-trabajo-en-equipo.png')
                    
                    new_image = image.resize((300, 99))
                    #html_msg = loader.render_to_string('C:/Users/gabri/OneDrive/Documentos/Repositorios-github/sistema_gestion_actividades/Sistema_gestion_actividades/task/templates/sendemail.html', {
                    html_msg = loader.render_to_string('/home/proyecto/sistema_gestion_actividades/Sistema_gestion_actividades/task/templates/sendemail.html', {
                        "Usuario": " ".join(list(map(lambda x: x.capitalize(), user.username.split(" ")))),
                        "tarea": str(task_day.description),
                        'fecha_inicio': str(task_day.start_day),
                        'fecha_entrega': str(task_day.end_day),
                    })
                    

                    html_msg = html_msg.replace('%% my_image %%', '{{ my_image }}')
                    email = EmailSender(
                        host=settings.EMAIL_HOST,
                        port=settings.EMAIL_PORT,
                        username=settings.EMAIL_HOST_USER,
                        password=settings.EMAIL_HOST_PASSWORD
                    )
                    email.send(
                        sender=settings.EMAIL_HOST_USER,
                        receivers=[user.email],
                        subject="Tarea Asignada",
                        html=html_msg,
                        body_images={"my_image": new_image}
                    )
                    print("Email enviado a: ", user.email)
            
            return Response({'access': (True), 'User': user.username}, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({'messages': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
    def create(self, request, *args, **kwargs):
        serialize = TaskSerializer(data=request.data)

        if serialize.is_valid():
            start_day = serialize.validated_data['start_day']
            end_day = serialize.validated_data['end_day']
            
            start_day_actual = datetime.date(datetime.now())
            print(start_day_actual)
            #replace_data = start_day_actual.replace(microsecond=0)
            if start_day < start_day_actual:
                return Response({'Messages': 'No puedes crear una tarea con una fecha del dia de ayer'}, status=status.HTTP_400_BAD_REQUEST)
            
            elif end_day <= start_day:
                return Response({'Messages': 'La fecha de entrega tiene que ser mayor o igual al dia que se asigno la fecha'}, status=status.HTTP_400_BAD_REQUEST)
            
            serialize.save()
            task_t = TaskListSerializer(instance=serialize.instance).data

            for user_id in task_t["user"]:
                self.send_email(request, task_t["id"], user_id)

            return Response(TaskListSerializer(instance=serialize.instance).data, status=status.HTTP_201_CREATED)
        else:
            return Response(serialize.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, pk=None):
        try:
            task_update = self.get_object()
            serializer = TaskListSerializer(task_update, data=request.data, partial=True)

            if serializer.is_valid(raise_exception=True):
                task_finish = serializer.validated_data.get('is_finished')
                porcentage_task = serializer.validated_data.get('porcentage_task')
                if task_finish == True and porcentage_task == 100:
                    serializer.save()

                    user_ids = request.data.get('user', [])
                    users = UserCustomer.objects.filter(is_active=True, id__in=user_ids)

                    for user in users:
                        history = TaskHistory(
                            task=task_update,
                            name=task_update.name,
                            description=task_update.description,
                            is_finished=task_update.is_finished,
                            departament=task_update.departament,
                            porcentage_task=task_update.porcentage_task,
                            start_day=task_update.start_day,
                            end_day=task_update.end_day,
                        )
                        history.save()
                        history.user.set([user.id])

                    return Response(TaskListSerializer(instance=serializer.instance).data, status=status.HTTP_200_OK)
                else:
                    serializer.save()
                    return Response(TaskListSerializer(instance=serializer.instance).data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({'messages': str(e)}, status=status.HTTP_400_BAD_REQUEST)

class ReportTaskFinished(APIView):
    queryset = TaskHistory.objects.all()
    serializer_class = TaskHistoryListSerializer
    filter_backends = (DjangoFilterBackend, CreatedBetweenFilter)
    filter_fields = ('is_enabled',)
    
    def get(self, request, *args, **kwargs):
        # Obtener los objetos de TaskHistory según los filtros
        queryset = TaskHistory.objects.filter(is_enabled=True).order_by('-created')

        # Serializar los datos que deseas devolver como JSON
        serializer = self.serializer_class(queryset, many=True)
        data = serializer.data

        # Crear un nuevo archivo de Excel
        wb = openpyxl.Workbook()
        sheet = wb.active
        users = UserCustomer.objects.all()
        user_names = ', '.join(user.username for user in users if user.username)

        # Crear los encabezados de las columnas
        headers = ['ID', 'name', 'description', 'is_enabled', 'is_finished', 'is_pending', 'is_started', 'project', 'user', 'departament', 'start_day', 'end_day']  # Reemplaza los campos con los correctos
        for col_num, header_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet.column_dimensions[col_letter].width = 15
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header_title
            cell.font = Font(bold=True)

        # Llenar la información de las filas
        for row_num, obj in enumerate(queryset, 2):
            sheet.cell(row=row_num, column=1).value = obj.id
            sheet.cell(row=row_num, column=2).value = obj.name
            sheet.cell(row=row_num, column=3).value = obj.description
            sheet.cell(row=row_num, column=4).value = obj.is_enabled
            sheet.cell(row=row_num, column=5).value = obj.is_finished
            sheet.cell(row=row_num, column=6).value = obj.is_pending
            sheet.cell(row=row_num, column=7).value = obj.is_started
            sheet.cell(row=row_num, column=8).value = obj.project
            sheet.cell(row=row_num, column=9).value = user_names
            # Opción alternativa: sheet.cell(row=row_num, column=8).value = obj.user.first().username  # Obtener el primer nombre de usuario
            sheet.cell(row=row_num, column=10).value = obj.departament.name
            sheet.cell(row=row_num, column=11).value = obj.start_day
            sheet.cell(row=row_num, column=12).value = obj.end_day

        # Guardar el archivo de Excel
        file_path = os.path.join(f'{settings.STATIC_ROOT}/media', 'task_finished.xlsx')
        wb.save(file_path)

        # Crear la respuesta HTTP con el archivo adjunto y los datos serializados
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="task_finished.xlsx"'
        with open(file_path, 'rb') as excel_file:
            response.write(excel_file.read())

        return Response(queryset, status=status.HTTP_200_OK)
        #return response
    

class project_tasksViewSet(ModelViewSet):
    #permission_classes = (IsAppAuthenticated, IsAppStaff, IsAuthenticated, IsSuperUser)
    serializer_class = project_tasks_list_serializer
    queryset = project_tasks.objects.filter(owner__rol__id=1)
    filter_backends = (DjangoFilterBackend,)
    filter_fields = ('id',)


    def create(self, request, *args, **kwargs):
        serialize = project_tasks_serializer(data=request.data)

        if serialize.is_valid():
            serialize.save()
            return Response(project_tasks_list_serializer(instance=serialize.instance).data, status=status.HTTP_200_OK)
        else:
            return Response(serialize.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        company = self.get_object()
        serialize = project_tasks_serializer(company, data=request.data, partial=True)

        if serialize.is_valid():
            serialize.save()

            return Response(project_tasks_list_serializer(instance=serialize.instance).data, status=status.HTTP_200_OK)
        else:
            return Response(serialize.errors, status=status.HTTP_400_BAD_REQUEST)